import cv2
import face_recognition
import numpy as np
import os,json
from datetime import datetime
from openpyxl import load_workbook, Workbook

# === CONFIGURATION ===
DATA_FILE = "data.xlsx"
IMAGE_PATH = "Images"
TODAY_FILE = f"D:/auto_attendance/output/{datetime.now().strftime('%Y-%m-%d')}.xlsx"

def load_time_slots():
    with open('time_slots.json', 'r') as f:
        data = json.load(f)
    return [tuple(slot) for slot in data]  # convert to list of tuples


TIME_SLOTS = load_time_slots()

# === FUNCTIONS ===
def load_students():
    if not os.path.exists(DATA_FILE):
        print("Error: data.xlsx not found!")
        exit()
    workbook = load_workbook(DATA_FILE)
    sheet = workbook.active
    students = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        s_no, roll_no, name, parent_no = row
        students[name.strip().upper()] = {"S/No": s_no, "Roll No": roll_no, "Parent No": parent_no}
    return students

students = load_students()

def load_known_faces():
    known_encodings = {}
    for file in os.listdir(IMAGE_PATH):
        if file.endswith(".jpg") or file.endswith(".png"):
            image_path = os.path.join(IMAGE_PATH, file)
            img = face_recognition.load_image_file(image_path)
            encoding = face_recognition.face_encodings(img)
            if encoding:
                name = os.path.splitext(file)[0].strip().upper()
                known_encodings[name] = encoding[0]
    return known_encodings

known_faces = load_known_faces()

if not os.path.exists(TODAY_FILE):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["S/No", "Student Name"] + [f"{start}-{end}" for start, end in TIME_SLOTS])
    for name, info in students.items():
        sheet.append([info["S/No"], name] + ["Absent"] * len(TIME_SLOTS))
    workbook.save(TODAY_FILE)

workbook = load_workbook(TODAY_FILE)
sheet = workbook.active
marked_names = {slot[0]: set() for slot in TIME_SLOTS}

def get_current_slot():
    now = datetime.now().strftime("%H:%M")
    now_time = datetime.strptime(now, "%H:%M")
    for start, end in TIME_SLOTS:
        start_time = datetime.strptime(start, "%H:%M")
        end_time = datetime.strptime(end, "%H:%M")
        if start_time <= now_time <= end_time:
            print(f"Active Slot Found: {start} - {end}")
            return start, end
    print(f"No active slot currently. Current Time: {now}")
    return None

def mark_attendance(name):
    current_slot = get_current_slot()
    if not current_slot:
        return
    start_time, end_time = current_slot
    if name in marked_names[start_time]:
        return
    name = name.strip().upper()

    # Auto-expand header if needed
    required_cols = len(TIME_SLOTS) + 2
    if sheet.max_column < required_cols:
        for idx, slot in enumerate(TIME_SLOTS):
            sheet.cell(row=1, column=idx + 3, value=f"{slot[0]}-{slot[1]}")
        workbook.save(TODAY_FILE)
        print(f"Auto-expanded columns to match {len(TIME_SLOTS)} time slots")

    # Find correct column
    try:
        slot_index = TIME_SLOTS.index((start_time, end_time))
    except ValueError:
        print("Current slot not found in TIME_SLOTS")
        return
    col_index = slot_index + 3  # Excel columns start from 1, +2 for name cols +1 for offset

    # Mark attendance
    for row in sheet.iter_rows(min_row=2, values_only=False):
        student_name = str(row[1].value).strip().upper()
        if student_name == name:
            row[col_index - 1].value = "Present"
            workbook.save(TODAY_FILE)
            marked_names[start_time].add(name)
            print(f"Marked {name} present for slot {start_time}-{end_time}")
            return
    print(f"{name} not found in the sheet")


# === PRE-PROCESSING FUNCTIONS ===
def adjust_brightness_contrast(image):
    alpha = 1.5  # Contrast
    beta = 50    # Brightness
    return cv2.convertScaleAbs(image, alpha=alpha, beta=beta)

def is_low_light(image):
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    mean_brightness = np.mean(gray)
    return mean_brightness < 50  # Tune this threshold if needed

# === FACE RECOGNITION SYSTEM ===
video_capture = cv2.VideoCapture(0)

while True:
    ret, frame = video_capture.read()
    if not ret:
        break

    # Low-light detection
    if is_low_light(frame):
        print(" Low light detected, skipping frame.")
        cv2.putText(frame, "Low Light - Please improve lighting!", (30, 50),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0, 0, 255), 2)
        cv2.imshow("Smart Attendance System", frame)
        if cv2.waitKey(1) & 0xFF == ord("q"):
            break
        continue

    # Pre-process image
    small_frame = cv2.resize(frame, (0, 0), fx=0.25, fy=0.25)
    bright_frame = adjust_brightness_contrast(small_frame)
    rgb_frame = cv2.cvtColor(bright_frame, cv2.COLOR_BGR2RGB)

    face_locations = face_recognition.face_locations(rgb_frame)
    face_encodings = face_recognition.face_encodings(rgb_frame, face_locations)

    for face_encoding, face_location in zip(face_encodings, face_locations):
        matches = face_recognition.compare_faces(list(known_faces.values()), face_encoding, tolerance=0.50)
        name = "Unknown"
        face_distances = face_recognition.face_distance(list(known_faces.values()), face_encoding)
        best_match_index = np.argmin(face_distances) if face_distances.size else None
        if best_match_index is not None and matches[best_match_index]:
            name = list(known_faces.keys())[best_match_index]
            mark_attendance(name)

        # Draw box
        top, right, bottom, left = [v * 4 for v in face_location]
        cv2.rectangle(frame, (left, top), (right, bottom), (0, 255, 0), 2)
        cv2.putText(frame, name, (left, top - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0, 255, 0), 2)

    cv2.imshow("Smart Attendance System", frame)

    if cv2.waitKey(1) & 0xFF == ord("q"):
        break

video_capture.release()
cv2.destroyAllWindows()
