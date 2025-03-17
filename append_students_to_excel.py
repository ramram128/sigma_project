import json
from openpyxl import load_workbook

def add_to_database():
    with open("students.json", "r") as f:
        students = json.load(f)

    workbook = load_workbook("data.xlsx")
    sheet = workbook.active

    # Get all roll numbers already in Excel to prevent duplicates
    existing_roll_numbers = set()
    for row in sheet.iter_rows(min_row=2, values_only=True):
        existing_roll_numbers.add(row[1])  # Assuming Roll No is column 2

    # Append new students only if roll number doesn't exist
    for student in students:
        roll_no = student.get("Roll No") or student.get("roll_no")
        if roll_no not in existing_roll_numbers:
            sheet.append([
                student.get("S/No") or student.get("s_no"),
                roll_no,
                student.get("Student Name") or student.get("student_name"),
                student.get("Parent No") or student.get("parent_no")
            ])

    workbook.save("data.xlsx")
    print("✅ Excel updated successfully!")

if __name__ == "__main__":
    add_to_database()
